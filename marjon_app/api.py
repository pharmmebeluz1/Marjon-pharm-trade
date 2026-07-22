from __future__ import annotations

import io
import json
from datetime import datetime, time, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from flask import Blueprint, current_app, jsonify, request, send_file
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import func, or_

from .extensions import db
from .models import (
    AuditLog,
    Branch,
    CourierLocation,
    HealthPassport,
    Order,
    OrderItem,
    PatientVault,
    Prescription,
    Product,
    Stock,
    User,
    utcnow,
)
from .security import audit, decrypt_json, encrypt_json, role_required
from .services import (
    ask_ai,
    can_transition,
    haversine_km,
    order_code,
    prescription_code,
    reserve_stock,
    restore_stock,
    save_private_upload,
)

bp = Blueprint("api", __name__, url_prefix="/api")


def _json() -> dict[str, Any]:
    value = request.get_json(silent=True)
    return value if isinstance(value, dict) else {}


def _parse_date(value: str | None, end: bool = False) -> datetime | None:
    if not value:
        return None
    try:
        day = datetime.strptime(value, "%Y-%m-%d").date()
        return datetime.combine(day, time.max if end else time.min, tzinfo=timezone.utc)
    except ValueError:
        return None


def _bounded_int(value: Any, default: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(parsed, maximum))


def _optional_coordinate(value: Any, minimum: float, maximum: float) -> float | None:
    if value in (None, ""):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        raise ValueError("Koordinata raqam bo‘lishi kerak.")
    if not minimum <= parsed <= maximum:
        raise ValueError("Koordinata ruxsat etilgan chegaradan tashqarida.")
    return parsed


def _accessible_order(order: Order) -> bool:
    if not current_user.is_authenticated:
        token = request.args.get("token", "")
        return bool(token and token == order.tracking_token)
    if current_user.role in {"manager", "admin", "pharmacist", "accountant"}:
        return True
    if current_user.role == "courier":
        return order.courier_id == current_user.id
    return order.patient_id == current_user.id


def _order_payload(order: Order, include_private: bool | None = None) -> dict[str, Any]:
    role = current_user.role if current_user.is_authenticated else ""
    if include_private is None:
        include_private = role in {"patient", "pharmacist", "courier", "manager", "admin"}
    data = order.to_dict(include_private=include_private)
    if role not in {"accountant", "manager", "admin"}:
        data.pop("cost", None)
        data.pop("profit", None)
    if role != "patient":
        data.pop("tracking_token", None)
    return data


def _order_query():
    query = Order.query.order_by(Order.created_at.desc())
    if current_user.role == "patient":
        query = query.filter(Order.patient_id == current_user.id)
    elif current_user.role == "courier":
        query = query.filter(Order.courier_id == current_user.id)
    start = _parse_date(request.args.get("start"))
    end = _parse_date(request.args.get("end"), end=True)
    if start:
        query = query.filter(Order.created_at >= start)
    if end:
        query = query.filter(Order.created_at <= end)
    branch = request.args.get("branch")
    if branch:
        query = query.join(Branch).filter((Branch.slug == branch) | (Branch.city == branch))
    status = request.args.get("status")
    if status:
        query = query.filter(Order.status == status)
    return query


@bp.get("/health")
def health():
    return jsonify({"ok": True, "service": "Pharm360°", "version": "8.0.0-secure"})


@bp.get("/products")
def products():
    query = Product.query.filter_by(is_active=True).order_by(Product.id)
    category = request.args.get("category")
    search = (request.args.get("q") or "").strip()
    if category and category != "all":
        query = query.filter(Product.category == category)
    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(Product.name_uz.ilike(like), Product.name_ru.ilike(like), Product.name_en.ilike(like), Product.sku.ilike(like))
        )
    return jsonify({"ok": True, "products": [product.to_dict() for product in query.all()]})


@bp.get("/branches")
def branches():
    rows = Branch.query.filter_by(is_active=True).order_by(Branch.id).all()
    return jsonify({"ok": True, "branches": [row.to_dict() for row in rows]})


@bp.get("/users")
@role_required("pharmacist", "manager", "admin")
def users():
    role = request.args.get("role", "")
    query = User.query.filter_by(is_active_account=True)
    if role:
        query = query.filter_by(role=role)
    return jsonify({"ok": True, "users": [user.to_dict() for user in query.order_by(User.name).all()]})


@bp.get("/orders")
@login_required
def list_orders():
    if current_user.must_change_password:
        return jsonify({"ok": False, "error": "Avval vaqtinchalik parolni almashtiring.", "code": "PASSWORD_CHANGE_REQUIRED"}), 403
    rows = _order_query().limit(_bounded_int(request.args.get("limit"), 200, 1, 500)).all()
    return jsonify({"ok": True, "orders": [_order_payload(row) for row in rows]})


@bp.post("/orders")
@role_required("patient")
def create_order():
    payload = _json()
    raw_items = payload.get("items")
    if not isinstance(raw_items, list) or not raw_items:
        return jsonify({"ok": False, "error": "Savat bo‘sh."}), 400
    if not bool(payload.get("consent")):
        return jsonify({"ok": False, "error": "Buyurtma ma’lumotlarini tasdiqlang."}), 400

    branch = None
    branch_id = payload.get("branch_id")
    branch_slug = str(payload.get("branch", "")).strip().lower()
    if branch_id:
        try:
            branch = db.session.get(Branch, int(branch_id))
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "Filial identifikatori noto‘g‘ri."}), 400
    if not branch and branch_slug:
        branch = Branch.query.filter((Branch.slug == branch_slug) | (func.lower(Branch.city) == branch_slug)).first()
    if not branch:
        branch = Branch.query.filter_by(is_active=True).order_by(Branch.id).first()
    if not branch:
        return jsonify({"ok": False, "error": "Faol filial topilmadi."}), 503

    items: list[OrderItem] = []
    total = Decimal("0")
    cost_total = Decimal("0")
    for item in raw_items[:100]:
        try:
            product_id = int(item.get("id"))
            quantity = int(item.get("qty", 1))
        except (TypeError, ValueError, AttributeError):
            return jsonify({"ok": False, "error": "Savatdagi mahsulot ma’lumoti noto‘g‘ri."}), 400
        if quantity < 1 or quantity > 50:
            return jsonify({"ok": False, "error": "Mahsulot miqdori 1–50 oralig‘ida bo‘lsin."}), 400
        product = db.session.get(Product, product_id)
        if not product or not product.is_active:
            return jsonify({"ok": False, "error": f"Mahsulot topilmadi: {product_id}"}), 404
        stock = Stock.query.filter_by(branch_id=branch.id, product_id=product.id).first()
        if not stock or stock.quantity < quantity:
            return jsonify({"ok": False, "error": f"{product.name_uz} uchun {branch.city} filialida qoldiq yetarli emas."}), 409
        line_total = Decimal(product.price) * quantity
        line_cost = Decimal(product.cost_price) * quantity
        total += line_total
        cost_total += line_cost
        items.append(
            OrderItem(
                product_id=product.id,
                product_name=product.name_uz,
                quantity=quantity,
                unit_price=product.price,
                unit_cost=product.cost_price,
            )
        )

    code = order_code()
    while Order.query.filter_by(code=code).first():
        code = order_code()
    method = str(payload.get("method", "delivery"))[:20]
    if method not in {"delivery", "pickup"}:
        method = "delivery"
    payment_method = str(payload.get("payment", "Naqd"))[:40]
    allowed_payments = {"Naqd", "Click", "Payme", "Uzcard / Humo", "Terminal", "Bank o‘tkazmasi", "Qarz"}
    if payment_method not in allowed_payments:
        return jsonify({"ok": False, "error": "To‘lov usuli noto‘g‘ri."}), 400
    address = str(payload.get("address", current_user.address or branch.address)).strip()[:500]
    if method == "delivery" and len(address) < 5:
        return jsonify({"ok": False, "error": "Yetkazib berish manzilini to‘liq yozing."}), 400
    try:
        latitude = _optional_coordinate(payload.get("latitude"), -90, 90)
        longitude = _optional_coordinate(payload.get("longitude"), -180, 180)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    order = Order(
        code=code,
        patient_id=current_user.id,
        branch_id=branch.id,
        method=method,
        payment_method=payment_method,
        payment_status="pending",
        status="Yangi",
        address=address,
        latitude=latitude,
        longitude=longitude,
        total=total,
        cost_total=cost_total,
        notes=str(payload.get("notes", ""))[:1000],
        items=items,
    )
    db.session.add(order)
    db.session.flush()
    audit("order_created", "order", order.id, f"code={order.code}; total={int(total)}")
    db.session.commit()
    return jsonify({"ok": True, "order": _order_payload(order), "message": "Buyurtma serverga saqlandi."}), 201


@bp.get("/orders/<string:code>")
@login_required
def get_order(code: str):
    order = Order.query.filter_by(code=code).first_or_404()
    if not _accessible_order(order):
        return jsonify({"ok": False, "error": "Bu buyurtmani ko‘rishga ruxsat yo‘q."}), 403
    return jsonify({"ok": True, "order": _order_payload(order)})


@bp.patch("/orders/<string:code>/status")
@role_required("pharmacist", "courier", "manager", "admin")
def update_order_status(code: str):
    order = Order.query.filter_by(code=code).first_or_404()
    if current_user.role == "courier" and order.courier_id != current_user.id:
        return jsonify({"ok": False, "error": "Bu buyurtma sizga biriktirilmagan."}), 403
    payload = _json()
    new_status = str(payload.get("status", "")).strip()
    if not can_transition(order, new_status, current_user.role):
        return jsonify({"ok": False, "error": f"{order.status} holatidan {new_status} holatiga o‘tishga ruxsat yo‘q."}), 409

    if new_status in {"Tayyorlanmoqda", "Tayyor", "Kuryerga berildi"}:
        ok, message = reserve_stock(order)
        if not ok:
            db.session.rollback()
            return jsonify({"ok": False, "error": message}), 409
    if new_status == "Bekor qilindi":
        restore_stock(order)
        order.cancelled_at = utcnow()
    if new_status == "Yetkazildi":
        order.delivered_at = utcnow()
        if order.payment_method == "Naqd":
            order.payment_status = "paid"

    old = order.status
    order.status = new_status
    audit("order_status_changed", "order", order.id, f"{old} -> {new_status}")
    db.session.commit()
    return jsonify({"ok": True, "order": _order_payload(order)})


@bp.patch("/orders/<string:code>/assign-courier")
@role_required("pharmacist", "manager", "admin")
def assign_courier(code: str):
    order = Order.query.filter_by(code=code).first_or_404()
    payload = _json()
    try:
        courier_id = int(payload.get("courier_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Kuryerni tanlang."}), 400
    courier = db.session.get(User, courier_id)
    if not courier or courier.role != "courier" or not courier.is_active_account:
        return jsonify({"ok": False, "error": "Faol kuryer topilmadi."}), 404
    order.courier_id = courier.id
    if order.status in {"Tayyor", "Tayyorlanmoqda"}:
        ok, message = reserve_stock(order)
        if not ok:
            db.session.rollback()
            return jsonify({"ok": False, "error": message}), 409
        order.status = "Kuryerga berildi"
    audit("courier_assigned", "order", order.id, f"courier={courier.name}")
    db.session.commit()
    return jsonify({"ok": True, "order": _order_payload(order)})


@bp.post("/orders/<string:code>/courier-location")
@role_required("courier", "manager", "admin")
def courier_location(code: str):
    order = Order.query.filter_by(code=code).first_or_404()
    if current_user.role == "courier" and order.courier_id != current_user.id:
        return jsonify({"ok": False, "error": "Bu buyurtma sizga biriktirilmagan."}), 403
    payload = _json()
    try:
        lat = float(payload["latitude"])
        lon = float(payload["longitude"])
    except (KeyError, TypeError, ValueError):
        return jsonify({"ok": False, "error": "GPS koordinatalari noto‘g‘ri."}), 400
    if not (-90 <= lat <= 90 and -180 <= lon <= 180):
        return jsonify({"ok": False, "error": "GPS koordinatalari chegaradan tashqarida."}), 400
    courier_id = order.courier_id or current_user.id
    if not order.courier_id and current_user.role == "courier":
        order.courier_id = current_user.id
    location = CourierLocation(
        order_id=order.id,
        courier_id=courier_id,
        latitude=lat,
        longitude=lon,
        accuracy=payload.get("accuracy"),
        heading=payload.get("heading"),
        speed=payload.get("speed"),
    )
    db.session.add(location)
    if order.status == "Kuryerga berildi":
        order.status = "Yetkazilmoqda"
    db.session.commit()
    return jsonify({"ok": True, "location": location.to_dict(), "order_status": order.status})


@bp.get("/orders/<string:code>/tracking")
def tracking(code: str):
    order = Order.query.filter_by(code=code).first_or_404()
    if not _accessible_order(order):
        return jsonify({"ok": False, "error": "Kuzatuv uchun ruxsat yo‘q."}), 403
    latest = CourierLocation.query.filter_by(order_id=order.id).order_by(CourierLocation.created_at.desc()).first()
    distance = None
    eta_minutes = None
    if latest and order.latitude is not None and order.longitude is not None:
        distance = haversine_km(latest.latitude, latest.longitude, order.latitude, order.longitude)
        eta_minutes = max(2, round(distance / 22 * 60))
    authenticated_access = current_user.is_authenticated and _accessible_order(order)
    order_data = _order_payload(order, include_private=authenticated_access)
    if not authenticated_access:
        # A shared tracking token reveals only delivery progress, never health, contact,
        # exact destination, payment or internal financial information.
        for key in ("cost", "profit", "payment", "payment_status", "tracking_token", "courier_phone"):
            order_data.pop(key, None)
    return jsonify(
        {
            "ok": True,
            "order": order_data,
            "location": latest.to_dict() if latest else None,
            "distance_km": round(distance, 2) if distance is not None else None,
            "eta_minutes": eta_minutes,
            "tracking_active": bool(latest and order.status in {"Kuryerga berildi", "Yetkazilmoqda"}),
        }
    )


@bp.get("/health-passport")
@role_required("patient")
def get_health_passport():
    record = HealthPassport.query.filter_by(user_id=current_user.id).first()
    return jsonify({"ok": True, "passport": decrypt_json(record.encrypted_payload) if record else {}})


@bp.put("/health-passport")
@role_required("patient")
def save_health_passport():
    payload = _json()
    if not bool(payload.get("consent")):
        return jsonify({"ok": False, "error": "Ma’lumotlarni saqlash uchun rozilikni tasdiqlang."}), 400
    clean = {
        "name": str(payload.get("name", ""))[:120],
        "birth": str(payload.get("birth", ""))[:20],
        "blood": str(payload.get("blood", ""))[:5],
        "phone": str(payload.get("phone", ""))[:30],
        "allergy": str(payload.get("allergy", ""))[:2000],
        "medicines": str(payload.get("medicines", ""))[:2000],
        "age": str(payload.get("age", ""))[:3],
        "emergency_name": str(payload.get("emergency_name", ""))[:120],
        "emergency_phone": str(payload.get("emergency_phone", ""))[:30],
        "updated_at": utcnow().isoformat(),
    }
    record = HealthPassport.query.filter_by(user_id=current_user.id).first()
    if not record:
        record = HealthPassport(user_id=current_user.id)
        db.session.add(record)
    record.encrypted_payload = encrypt_json(clean)
    record.consent = True
    audit("health_passport_saved", "health_passport", current_user.id)
    db.session.commit()
    return jsonify({"ok": True, "passport": clean})


@bp.delete("/health-passport")
@role_required("patient")
def delete_health_passport():
    record = HealthPassport.query.filter_by(user_id=current_user.id).first()
    if record:
        db.session.delete(record)
        audit("health_passport_deleted", "health_passport", current_user.id)
        db.session.commit()
    return jsonify({"ok": True})


def _clean_patient_vault(payload: dict[str, Any]) -> dict[str, Any]:
    reminders: list[dict[str, Any]] = []
    for item in payload.get("reminders", [])[:50] if isinstance(payload.get("reminders"), list) else []:
        if not isinstance(item, dict):
            continue
        medicine = str(item.get("medicine", "")).strip()[:180]
        reminder_time = str(item.get("time", "")).strip()[:10]
        if medicine and reminder_time:
            reminders.append({
                "id": _bounded_int(item.get("id"), 0, 0, 9_999_999_999_999),
                "medicine": medicine,
                "time": reminder_time,
                "active": bool(item.get("active", True)),
            })

    family_members: list[dict[str, str]] = []
    for item in payload.get("family_members", [])[:30] if isinstance(payload.get("family_members"), list) else []:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "")).strip()[:120]
        relation = str(item.get("relation", "")).strip()[:80]
        if name:
            family_members.append({"name": name, "relation": relation})

    reservations: list[dict[str, Any]] = []
    for item in payload.get("reservations", [])[:30] if isinstance(payload.get("reservations"), list) else []:
        if not isinstance(item, dict):
            continue
        code = str(item.get("code", "")).strip()[:40]
        product = str(item.get("product", "")).strip()[:180]
        branch = str(item.get("branch", "")).strip()[:120]
        expires = _bounded_int(item.get("expires"), 0, 0, 9_999_999_999_999)
        if code and product:
            reservations.append({"code": code, "product": product, "branch": branch, "expires": expires})

    return {"reminders": reminders, "family_members": family_members, "reservations": reservations}


@bp.get("/patient-vault")
@role_required("patient")
def get_patient_vault():
    record = PatientVault.query.filter_by(user_id=current_user.id).first()
    vault = decrypt_json(record.encrypted_payload) if record else {}
    return jsonify({"ok": True, "vault": _clean_patient_vault(vault)})


@bp.put("/patient-vault")
@role_required("patient")
def save_patient_vault():
    clean = _clean_patient_vault(_json())
    record = PatientVault.query.filter_by(user_id=current_user.id).first()
    if not record:
        record = PatientVault(user_id=current_user.id)
        db.session.add(record)
    record.encrypted_payload = encrypt_json(clean)
    audit("patient_vault_saved", "patient_vault", current_user.id)
    db.session.commit()
    return jsonify({"ok": True, "vault": clean})


@bp.delete("/patient-vault")
@role_required("patient")
def delete_patient_vault():
    record = PatientVault.query.filter_by(user_id=current_user.id).first()
    if record:
        db.session.delete(record)
        audit("patient_vault_deleted", "patient_vault", current_user.id)
        db.session.commit()
    return jsonify({"ok": True})


@bp.post("/prescriptions")
@role_required("patient")
def create_prescription():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"ok": False, "error": "Retsept faylini tanlang."}), 400
    if request.form.get("consent") not in {"true", "1", "on", "yes"}:
        return jsonify({"ok": False, "error": "Farmatsevt tekshiruvi uchun rozilikni tasdiqlang."}), 400
    branch = None
    branch_value = request.form.get("branch", "")
    if branch_value.isdigit():
        branch = db.session.get(Branch, int(branch_value))
    if not branch:
        branch = Branch.query.filter((Branch.slug == branch_value.lower()) | (Branch.city == branch_value)).first()
    if not branch:
        branch = Branch.query.first()
    try:
        original, stored, mime_type = save_private_upload(upload)
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    code = prescription_code()
    while Prescription.query.filter_by(code=code).first():
        code = prescription_code()
    record = Prescription(
        code=code,
        patient_id=current_user.id,
        branch_id=branch.id,
        original_filename=original,
        stored_filename=stored,
        mime_type=mime_type,
        status="Yangi",
        consent=True,
    )
    db.session.add(record)
    db.session.flush()
    audit("prescription_uploaded", "prescription", record.id, f"code={code}; file={original}")
    db.session.commit()
    return jsonify({"ok": True, "prescription": record.to_dict()}), 201


@bp.get("/prescriptions")
@login_required
def list_prescriptions():
    if current_user.role == "patient":
        rows = Prescription.query.filter_by(patient_id=current_user.id).order_by(Prescription.created_at.desc()).all()
    elif current_user.role in {"pharmacist", "manager", "admin"}:
        rows = Prescription.query.order_by(Prescription.created_at.desc()).limit(200).all()
    else:
        return jsonify({"ok": False, "error": "Bu bo‘lim uchun ruxsat yo‘q."}), 403
    return jsonify({"ok": True, "prescriptions": [row.to_dict() for row in rows]})


@bp.patch("/prescriptions/<int:prescription_id>")
@role_required("pharmacist", "manager", "admin")
def update_prescription(prescription_id: int):
    record = db.session.get(Prescription, prescription_id)
    if not record:
        return jsonify({"ok": False, "error": "Retsept topilmadi."}), 404
    payload = _json()
    status = str(payload.get("status", record.status))[:30]
    if status not in {"Yangi", "Tekshirilmoqda", "Tasdiqlandi", "Rad etildi", "Bajarildi"}:
        return jsonify({"ok": False, "error": "Retsept holati noto‘g‘ri."}), 400
    record.status = status
    record.pharmacist_id = current_user.id
    record.pharmacist_note = str(payload.get("note", record.pharmacist_note))[:1000]
    audit("prescription_updated", "prescription", record.id, f"status={status}")
    db.session.commit()
    return jsonify({"ok": True, "prescription": record.to_dict()})


@bp.get("/prescriptions/<int:prescription_id>/file")
@login_required
def prescription_file(prescription_id: int):
    record = db.session.get(Prescription, prescription_id)
    if not record:
        return jsonify({"ok": False, "error": "Retsept topilmadi."}), 404
    if current_user.role == "patient" and record.patient_id != current_user.id:
        return jsonify({"ok": False, "error": "Ruxsat yo‘q."}), 403
    if current_user.role not in {"patient", "pharmacist", "manager", "admin"}:
        return jsonify({"ok": False, "error": "Ruxsat yo‘q."}), 403
    path = Path(current_app.config["UPLOAD_FOLDER"]) / record.stored_filename
    if not path.exists():
        return jsonify({"ok": False, "error": "Fayl serverda topilmadi."}), 404
    audit("prescription_file_viewed", "prescription", record.id)
    db.session.commit()
    return send_file(path, mimetype=record.mime_type, download_name=record.original_filename, as_attachment=False)


@bp.get("/stock")
@role_required("pharmacist", "accountant", "manager", "admin")
def stock():
    rows = Stock.query.join(Product).join(Branch).order_by(Product.name_uz, Branch.id).all()
    return jsonify(
        {
            "ok": True,
            "stock": [
                {
                    "id": row.id,
                    "product_id": row.product_id,
                    "product": row.product.name_uz,
                    "branch_id": row.branch_id,
                    "branch": row.branch.city,
                    "quantity": row.quantity,
                    "minimum_quantity": row.minimum_quantity,
                    "low": row.quantity <= row.minimum_quantity,
                }
                for row in rows
            ],
        }
    )


@bp.patch("/stock/<int:stock_id>")
@role_required("pharmacist", "manager", "admin")
def update_stock(stock_id: int):
    row = db.session.get(Stock, stock_id)
    if not row:
        return jsonify({"ok": False, "error": "Ombor yozuvi topilmadi."}), 404
    payload = _json()
    try:
        quantity = int(payload.get("quantity", row.quantity))
        minimum = int(payload.get("minimum_quantity", row.minimum_quantity))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Qoldiq son bilan yozilishi kerak."}), 400
    if quantity < 0 or minimum < 0:
        return jsonify({"ok": False, "error": "Qoldiq manfiy bo‘lishi mumkin emas."}), 400
    old = row.quantity
    row.quantity = quantity
    row.minimum_quantity = minimum
    audit("stock_updated", "stock", row.id, f"{old} -> {quantity}")
    db.session.commit()
    return jsonify({"ok": True, "stock": {"id": row.id, "quantity": row.quantity, "minimum_quantity": row.minimum_quantity}})


@bp.get("/dashboard")
@role_required("accountant", "manager", "admin")
def dashboard():
    orders = _order_query().all()
    sales = sum(int(order.total or 0) for order in orders if order.status != "Bekor qilindi")
    cost = sum(int(order.cost_total or 0) for order in orders if order.status != "Bekor qilindi")
    low_stock = Stock.query.filter(Stock.quantity <= Stock.minimum_quantity).count()
    active_delivery = Order.query.filter(Order.status.in_(["Kuryerga berildi", "Yetkazilmoqda"])).count()
    return jsonify(
        {
            "ok": True,
            "stats": {
                "sales": sales,
                "cost": cost,
                "profit": sales - cost,
                "orders": len(orders),
                "low_stock": low_stock,
                "active_delivery": active_delivery,
            },
        }
    )


@bp.get("/audit")
@role_required("manager", "admin")
def audit_logs():
    rows = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(300).all()
    return jsonify({"ok": True, "logs": [row.to_dict() for row in rows]})


@bp.get("/reports/orders.xlsx")
@role_required("accountant", "manager", "admin")
def export_orders_xlsx():
    rows = _order_query().all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Buyurtmalar"
    headers = ["Kod", "Sana", "Bemor", "Telefon", "Filial", "Dorilar", "Jami", "Tannarx", "Foyda", "To‘lov", "To‘lov holati", "Holat", "Manzil", "Kuryer"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for order in rows:
        sheet.append(
            [
                order.code,
                order.created_at.astimezone().strftime("%d.%m.%Y %H:%M"),
                order.patient.name,
                order.patient.phone,
                order.branch.city,
                ", ".join(f"{item.product_name} × {item.quantity}" for item in order.items),
                int(order.total or 0),
                int(order.cost_total or 0),
                int(order.profit),
                order.payment_method,
                order.payment_status,
                order.status,
                order.address,
                order.courier.name if order.courier else "",
            ]
        )
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 55)

    payment_sheet = workbook.create_sheet("To‘lovlar")
    payment_sheet.append(["To‘lov turi", "Jami"])
    payments: dict[str, int] = {}
    for order in rows:
        if order.status == "Bekor qilindi":
            continue
        payments[order.payment_method] = payments.get(order.payment_method, 0) + int(order.total or 0)
    for method, total in sorted(payments.items()):
        payment_sheet.append([method, total])

    stock_sheet = workbook.create_sheet("Ombor")
    stock_sheet.append(["Mahsulot", "Filial", "Qoldiq", "Minimal qoldiq", "Holat"])
    for row in Stock.query.join(Product).join(Branch).order_by(Product.name_uz, Branch.id).all():
        stock_sheet.append([row.product.name_uz, row.branch.city, row.quantity, row.minimum_quantity, "Kam" if row.quantity <= row.minimum_quantity else "Yetarli"])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    audit("excel_exported", "report", "orders", f"rows={len(rows)}")
    db.session.commit()
    filename = f"Pharm360_Hisobot_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.post("/ai/ask")
def ai_ask():
    payload = _json()
    question = str(payload.get("question", ""))[:2000]
    if not question.strip():
        return jsonify({"ok": False, "error": "Savolni yozing."}), 400
    language = str(payload.get("language", "uz"))[:5]
    result = ask_ai(question, language)
    return jsonify({"ok": True, **result, "disclaimer": "AI tashxis qo‘ymaydi va dori tayinlamaydi."})


@bp.post("/payments/create")
@role_required("patient")
def create_payment():
    payload = _json()
    order = Order.query.filter_by(code=str(payload.get("order_code", "")), patient_id=current_user.id).first()
    if not order:
        return jsonify({"ok": False, "error": "Buyurtma topilmadi."}), 404
    provider = str(payload.get("provider", order.payment_method)).lower()
    if provider not in {"click", "payme"}:
        return jsonify({"ok": False, "error": "Onlayn to‘lov uchun Click yoki Payme tanlang."}), 400
    return jsonify(
        {
            "ok": False,
            "configured": False,
            "error": f"{provider.title()} merchant kalitlari hali ulanmagan. Buyurtma saqlandi, to‘lov holati pending.",
            "order": _order_payload(order),
        }
    ), 501
