from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import load_workbook

from marjon_app import create_app
from marjon_app.extensions import db
from marjon_app.models import HealthPassport, Order, PatientVault


@pytest.fixture()
def app(tmp_path: Path):
    database = tmp_path / "test.db"
    uploads = tmp_path / "uploads"
    application = create_app(
        {
            "TESTING": True,
            "APP_ENV": "testing",
            "SECRET_KEY": "test-secret-only",
            "DATA_ENCRYPTION_KEY": "",
            "SQLALCHEMY_DATABASE_URI": f"sqlite:///{database}",
            "UPLOAD_FOLDER": str(uploads),
            "WTF_CSRF_ENABLED": False,
            "REQUIRE_SMS_OTP": False,
        }
    )
    yield application
    with application.app_context():
        db.session.remove()
        db.drop_all()


@pytest.fixture()
def client(app):
    return app.test_client()


def csrf(client) -> str:
    return client.get("/api/auth/csrf").get_json()["csrf_token"]


def post(client, url: str, payload: dict, token: str | None = None):
    return client.post(url, json=payload, headers={"X-CSRF-Token": token or csrf(client)})


def patch(client, url: str, payload: dict, token: str | None = None):
    return client.patch(url, json=payload, headers={"X-CSRF-Token": token or csrf(client)})


def login(client, phone: str, password: str):
    response = post(client, "/api/auth/login", {"phone": phone, "password": password})
    assert response.status_code == 200, response.get_json()
    return response.get_json()


def change_password(client, current_password: str, new_password: str):
    response = post(
        client,
        "/api/auth/change-password",
        {"current_password": current_password, "new_password": new_password},
    )
    assert response.status_code == 200, response.get_json()


def logout(client):
    response = post(client, "/api/auth/logout", {})
    assert response.status_code == 200


def register_patient(client):
    response = post(
        client,
        "/api/auth/register",
        {
            "name": "Test Bemor",
            "phone": "+998901234567",
            "password": "Bemor2026!",
            "address": "Toshkent shahri, Chilonzor tumani",
            "consent": True,
        },
    )
    assert response.status_code == 201, response.get_json()
    return response.get_json()


def test_full_secure_order_and_tracking_flow(app, client):
    assert client.get("/").status_code == 200
    products = client.get("/api/products?q=para").get_json()["products"]
    assert products and products[0]["price"] > 0

    register_patient(client)
    token = csrf(client)
    response = post(
        client,
        "/api/orders",
        {
            "items": [{"id": products[0]["id"], "qty": 1}],
            "address": "Toshkent shahri, Chilonzor tumani, 10-uy",
            "method": "delivery",
            "payment": "Naqd",
            "latitude": 41.285,
            "longitude": 69.205,
            "consent": True,
        },
        token,
    )
    assert response.status_code == 201, response.get_json()
    order = response.get_json()["order"]
    code = order["code"]
    tracking_token = order["tracking_token"]
    assert "cost" not in order
    assert "profit" not in order
    with app.app_context():
        stored_order = Order.query.filter_by(code=code).first()
        assert int(stored_order.cost_total) < int(stored_order.total)

    response = client.put(
        "/api/health-passport",
        json={"name": "Test Bemor", "allergy": "Penitsillin", "consent": True},
        headers={"X-CSRF-Token": csrf(client)},
    )
    assert response.status_code == 200
    with app.app_context():
        raw = HealthPassport.query.first().encrypted_payload
        assert "Penitsillin" not in raw

    response = client.put(
        "/api/patient-vault",
        json={
            "reminders": [{"id": 1, "medicine": "Vitamin D", "time": "09:00", "active": True}],
            "family_members": [{"name": "Farzand", "relation": "Qizi"}],
            "reservations": [],
        },
        headers={"X-CSRF-Token": csrf(client)},
    )
    assert response.status_code == 200
    with app.app_context():
        raw_vault = PatientVault.query.first().encrypted_payload
        assert "Vitamin D" not in raw_vault
        assert "Farzand" not in raw_vault

    logout(client)
    public = client.get(f"/api/orders/{code}/tracking?token={tracking_token}")
    assert public.status_code == 200
    public_order = public.get_json()["order"]
    assert "phone" not in public_order
    assert "address" not in public_order
    assert "cost" not in public_order
    assert "profit" not in public_order

    login(client, "+998900000001", "Farm2026!")
    change_password(client, "Farm2026!", "FarmYangi2026!")
    for status in ["Tasdiqlandi", "Tayyorlanmoqda", "Tayyor"]:
        response = patch(client, f"/api/orders/{code}/status", {"status": status})
        assert response.status_code == 200, response.get_json()
    couriers = client.get("/api/users?role=courier").get_json()["users"]
    response = patch(client, f"/api/orders/{code}/assign-courier", {"courier_id": couriers[0]["id"]})
    assert response.status_code == 200
    logout(client)

    login(client, "+998900000002", "Kuryer2026!")
    change_password(client, "Kuryer2026!", "KuryerYangi2026!")
    response = post(
        client,
        f"/api/orders/{code}/courier-location",
        {"latitude": 41.30, "longitude": 69.22, "accuracy": 8, "speed": 7},
    )
    assert response.status_code == 200, response.get_json()
    live = client.get(f"/api/orders/{code}/tracking").get_json()
    assert live["tracking_active"] is True
    assert live["distance_km"] is not None
    logout(client)

    login(client, "+998900000000", "Marjon2026!")
    change_password(client, "Marjon2026!", "RahbarYangi2026!")
    dashboard = client.get("/api/dashboard").get_json()
    assert dashboard["ok"] is True
    report = client.get("/api/reports/orders.xlsx")
    assert report.status_code == 200
    workbook = load_workbook(io.BytesIO(report.data))
    assert "Buyurtmalar" in workbook.sheetnames


def test_validation_and_role_isolation(client):
    register_patient(client)
    product = client.get("/api/products").get_json()["products"][0]
    bad = post(
        client,
        "/api/orders",
        {
            "items": [{"id": product["id"], "qty": 1}],
            "address": "x",
            "method": "delivery",
            "payment": "Noma’lum",
            "consent": True,
        },
    )
    assert bad.status_code == 400
    assert client.get("/api/dashboard").status_code == 403

    invalid_limit = client.get("/api/orders?limit=noto‘g‘ri")
    assert invalid_limit.status_code == 200
